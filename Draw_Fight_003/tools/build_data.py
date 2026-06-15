from __future__ import annotations

import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
DATA_DIR = ROOT / "data"


def effect(items):
    return json.dumps(items, ensure_ascii=False)


def deck(items):
    return json.dumps(items, ensure_ascii=False)


KEYWORDS = [
    {"KeywordID": "KW_PENDING", "Name": "待封装", "Category": "区域", "Description": "编入后的牌进入待封装区，不会立即触发效果。"},
    {"KeywordID": "KW_CHAIN", "Name": "牌链", "Category": "区域", "Description": "封装后生成的自动循环序列，按牌序逐张触发。"},
    {"KeywordID": "KW_PLAYTIME", "Name": "打牌时间", "Category": "时间", "Description": "编入牌和牌链触发该牌时消耗的时间。"},
    {"KeywordID": "KW_BURN", "Name": "燃烧", "Category": "状态", "Description": "持续火焰资源，可被助燃和爆燃利用。每次自然结算造成层数伤害并减少1层。"},
    {"KeywordID": "KW_POISON", "Name": "中毒", "Category": "状态", "Description": "持续毒性资源，每次自然结算造成层数伤害并减少1层。"},
    {"KeywordID": "KW_BLOCK", "Name": "护盾", "Category": "状态", "Description": "抵消即将受到的伤害。"},
    {"KeywordID": "KW_FOCUS", "Name": "专注", "Category": "状态", "Description": "下一次造成的直接伤害提高，每次伤害后消耗1层。"},
    {"KeywordID": "KW_BLEED", "Name": "裂伤", "Category": "状态", "Description": "被刀刃体系利用的持续伤害资源。"},
]


CARDS = [
    {"CardID": "C001", "Name": "火花", "Archetype": "燃烧", "Rarity": "基础", "PlayTimeT": 0.5, "Tags": "伤害;燃烧", "Description": "造成3点伤害。", "Effects": effect([{"op": "damage", "amount": 3}])},
    {"CardID": "C002", "Name": "点燃", "Archetype": "燃烧", "Rarity": "基础", "PlayTimeT": 0.8, "Tags": "燃烧", "Description": "施加燃烧2。", "Effects": effect([{"op": "status", "target": "enemy", "status": "burn", "amount": 2}])},
    {"CardID": "C003", "Name": "助燃", "Archetype": "燃烧", "Rarity": "普通", "PlayTimeT": 1.0, "Tags": "燃烧;伤害", "Description": "敌人每有1层燃烧，造成1点伤害。", "Effects": effect([{"op": "burn_damage", "per": 1}])},
    {"CardID": "C004", "Name": "爆燃", "Archetype": "燃烧", "Rarity": "稀有", "PlayTimeT": 1.4, "Tags": "燃烧;爆发", "Description": "引爆燃烧，每层造成2点伤害，然后清除燃烧。", "Effects": effect([{"op": "ignite", "per": 2}])},
    {"CardID": "C005", "Name": "余烬护身", "Archetype": "燃烧", "Rarity": "普通", "PlayTimeT": 0.9, "Tags": "护盾;燃烧", "Description": "获得护盾4；若敌人有燃烧，额外获得护盾3。", "Effects": effect([{"op": "block", "amount": 4}, {"op": "if_enemy_status", "status": "burn", "op2": "block", "amount": 3}])},
    {"CardID": "C006", "Name": "格挡", "Archetype": "防御", "Rarity": "基础", "PlayTimeT": 0.7, "Tags": "护盾", "Description": "获得护盾6。", "Effects": effect([{"op": "block", "amount": 6}])},
    {"CardID": "C007", "Name": "修整", "Archetype": "防御", "Rarity": "基础", "PlayTimeT": 1.0, "Tags": "治疗;护盾", "Description": "回复3点生命；若有护盾，额外回复2点。", "Effects": effect([{"op": "heal", "amount": 3}, {"op": "if_self_status", "status": "block", "op2": "heal", "amount": 2}])},
    {"CardID": "C008", "Name": "反击预备", "Archetype": "防御", "Rarity": "普通", "PlayTimeT": 0.9, "Tags": "护盾;伤害", "Description": "获得护盾4；造成自身护盾25%的伤害。", "Effects": effect([{"op": "block", "amount": 4}, {"op": "block_strike", "ratio": 0.25}])},
    {"CardID": "C009", "Name": "铁壁", "Archetype": "防御", "Rarity": "稀有", "PlayTimeT": 1.3, "Tags": "护盾", "Description": "获得护盾12。", "Effects": effect([{"op": "block", "amount": 12}])},
    {"CardID": "C010", "Name": "短刺", "Archetype": "裂伤", "Rarity": "基础", "PlayTimeT": 0.45, "Tags": "伤害;裂伤", "Description": "造成2点伤害，施加裂伤1。", "Effects": effect([{"op": "damage", "amount": 2}, {"op": "status", "target": "enemy", "status": "bleed", "amount": 1}])},
    {"CardID": "C011", "Name": "割裂", "Archetype": "裂伤", "Rarity": "普通", "PlayTimeT": 0.9, "Tags": "裂伤", "Description": "施加裂伤3。", "Effects": effect([{"op": "status", "target": "enemy", "status": "bleed", "amount": 3}])},
    {"CardID": "C012", "Name": "撕开伤口", "Archetype": "裂伤", "Rarity": "稀有", "PlayTimeT": 1.2, "Tags": "裂伤;伤害", "Description": "敌人每有1层裂伤，造成2点伤害。", "Effects": effect([{"op": "bleed_damage", "per": 2}])},
    {"CardID": "C013", "Name": "毒针", "Archetype": "中毒", "Rarity": "基础", "PlayTimeT": 0.6, "Tags": "中毒;伤害", "Description": "造成2点伤害，施加中毒2。", "Effects": effect([{"op": "damage", "amount": 2}, {"op": "status", "target": "enemy", "status": "poison", "amount": 2}])},
    {"CardID": "C014", "Name": "毒雾", "Archetype": "中毒", "Rarity": "普通", "PlayTimeT": 1.1, "Tags": "中毒", "Description": "施加中毒4。", "Effects": effect([{"op": "status", "target": "enemy", "status": "poison", "amount": 4}])},
    {"CardID": "C015", "Name": "毒爆", "Archetype": "中毒", "Rarity": "稀有", "PlayTimeT": 1.5, "Tags": "中毒;爆发", "Description": "引爆中毒，每层造成2点伤害，然后清除中毒。", "Effects": effect([{"op": "poison_burst", "per": 2}])},
    {"CardID": "C016", "Name": "凝神", "Archetype": "专注", "Rarity": "基础", "PlayTimeT": 0.8, "Tags": "专注", "Description": "获得专注2。", "Effects": effect([{"op": "status", "target": "self", "status": "focus", "amount": 2}])},
    {"CardID": "C017", "Name": "精准射击", "Archetype": "专注", "Rarity": "普通", "PlayTimeT": 1.0, "Tags": "专注;伤害", "Description": "造成6点伤害；若有专注，额外造成4点伤害。", "Effects": effect([{"op": "damage", "amount": 6}, {"op": "if_self_status", "status": "focus", "op2": "damage", "amount": 4}])},
    {"CardID": "C018", "Name": "冷静补给", "Archetype": "专注", "Rarity": "普通", "PlayTimeT": 0.9, "Tags": "治疗;专注", "Description": "回复2点生命，获得专注1。", "Effects": effect([{"op": "heal", "amount": 2}, {"op": "status", "target": "self", "status": "focus", "amount": 1}])},
    {"CardID": "C019", "Name": "快装", "Archetype": "节奏", "Rarity": "普通", "PlayTimeT": 0.35, "Tags": "节奏", "Description": "造成1点伤害。短打牌时间，适合快速封装短链。", "Effects": effect([{"op": "damage", "amount": 1}])},
    {"CardID": "C020", "Name": "重击", "Archetype": "通用", "Rarity": "普通", "PlayTimeT": 1.3, "Tags": "伤害", "Description": "造成10点伤害。", "Effects": effect([{"op": "damage", "amount": 10}])},
    {"CardID": "C021", "Name": "战场急救", "Archetype": "通用", "Rarity": "普通", "PlayTimeT": 1.2, "Tags": "治疗", "Description": "回复7点生命。", "Effects": effect([{"op": "heal", "amount": 7}])},
    {"CardID": "C022", "Name": "清醒", "Archetype": "通用", "Rarity": "普通", "PlayTimeT": 0.8, "Tags": "净化;治疗", "Description": "清除自身燃烧、中毒、裂伤，回复2点生命。", "Effects": effect([{"op": "cleanse"}, {"op": "heal", "amount": 2}])},
    {"CardID": "C023", "Name": "护火循环", "Archetype": "燃烧", "Rarity": "稀有", "PlayTimeT": 1.1, "Tags": "护盾;燃烧", "Description": "获得护盾5；敌人有燃烧时造成5点伤害。", "Effects": effect([{"op": "block", "amount": 5}, {"op": "if_enemy_status", "status": "burn", "op2": "damage", "amount": 5}])},
    {"CardID": "C024", "Name": "腐蚀防线", "Archetype": "中毒", "Rarity": "普通", "PlayTimeT": 1.0, "Tags": "护盾;中毒", "Description": "获得护盾5，施加中毒1。", "Effects": effect([{"op": "block", "amount": 5}, {"op": "status", "target": "enemy", "status": "poison", "amount": 1}])},
]


CHARACTERS = [
    {"CharacterID": "R001", "Name": "燃烧工匠", "Stage": "成型", "HP": 46, "AIStyle": "燃烧", "DeckList": deck({"C001": 3, "C002": 3, "C003": 2, "C004": 1, "C005": 1, "C006": 2, "C007": 1}), "Description": "围绕火花、点燃、助燃和爆燃构造输出链。"},
    {"CharacterID": "R002", "Name": "盾链守卫", "Stage": "成型", "HP": 60, "AIStyle": "防御", "DeckList": deck({"C006": 4, "C007": 2, "C008": 3, "C009": 1, "C020": 1, "C021": 1}), "Description": "通过护盾链和反击预备形成稳定生存。"},
    {"CharacterID": "R003", "Name": "裂伤游侠", "Stage": "成型", "HP": 42, "AIStyle": "裂伤", "DeckList": deck({"C010": 4, "C011": 3, "C012": 2, "C016": 1, "C017": 2, "C019": 2}), "Description": "用短打牌时间建立高频裂伤链。"},
    {"CharacterID": "R004", "Name": "毒雾术士", "Stage": "成型", "HP": 44, "AIStyle": "中毒", "DeckList": deck({"C013": 3, "C014": 3, "C015": 2, "C018": 2, "C022": 1, "C024": 2}), "Description": "中毒堆叠与毒爆形成长周期爆发。"},
    {"CharacterID": "R005", "Name": "专注射手", "Stage": "发育前", "HP": 38, "AIStyle": "专注", "DeckList": deck({"C016": 3, "C017": 3, "C018": 2, "C019": 2, "C006": 2}), "Description": "低复杂度专注输出，适合观察短链。"},
    {"CharacterID": "R006", "Name": "混合学徒", "Stage": "发育前", "HP": 40, "AIStyle": "混合", "DeckList": deck({"C001": 2, "C002": 1, "C006": 2, "C010": 2, "C013": 2, "C019": 2, "C021": 1}), "Description": "混合牌堆，容易出现是否污染待封装区的选择。"},
]


CONFIG = [
    {"Key": "DrawTimeT", "Value": 1.0, "Description": "基础抽牌时间。"},
    {"Key": "ShuffleTimeT", "Value": 1.0, "Description": "洗牌时间。"},
    {"Key": "DecisionAuto", "Value": "编入", "Description": "决策窗口结束时默认行为。"},
    {"Key": "SealTimeT", "Value": 0.0, "Description": "MVP 暂不设置封装耗时。"},
    {"Key": "StatusTickT", "Value": 2.5, "Description": "状态自然结算间隔。"},
]


def parse_json_columns(row):
    parsed = dict(row)
    for key in ("Effects", "DeckList"):
        if key in parsed:
            parsed[key] = json.loads(parsed[key])
    return parsed


def write_workbook(filename, rows, sheet_name):
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    headers = list(rows[0].keys())
    ws.append(headers)
    for row in rows:
        ws.append([row.get(h, "") for h in headers])

    header_fill = PatternFill("solid", fgColor="273241")
    for cell in ws[1]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for column_cells in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in column_cells)
        width = min(max(max_len + 2, 10), 54)
        ws.column_dimensions[get_column_letter(column_cells[0].column)].width = width

    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    wb.save(DATA_DIR / filename)


def write_data_js():
    payload = {
        "keywords": KEYWORDS,
        "cards": [parse_json_columns(card) for card in CARDS],
        "characters": [parse_json_columns(character) for character in CHARACTERS],
        "config": CONFIG,
    }
    js = "window.BATTLE_DATA = " + json.dumps(payload, ensure_ascii=False, indent=2) + ";\n"
    (DATA_DIR / "game_data.js").write_text(js, encoding="utf-8")


def main():
    write_workbook("cards.xlsx", CARDS, "Cards")
    write_workbook("characters.xlsx", CHARACTERS, "Characters")
    write_workbook("keywords.xlsx", KEYWORDS, "Keywords")
    write_workbook("battle_config.xlsx", CONFIG, "BattleConfig")
    write_data_js()


if __name__ == "__main__":
    main()
